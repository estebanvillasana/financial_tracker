import os
import sqlite3
import tempfile
import time
from datetime import datetime
from decimal import Decimal

from fastapi import APIRouter
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel
from starlette.background import BackgroundTask

from database import get_connection, get_current_db_path
from scripts.backup_db import BACKUP_DIR, get_latest_backup_time

router = APIRouter(prefix="/app-config", tags=["App Config"])


def _read_currency() -> str:
    with get_connection() as conn:
        row = conn.execute(
            "SELECT value FROM settings WHERE key = 'currency'"
        ).fetchone()
        return row[0] if row else "usd"


def _write_currency(currency: str) -> None:
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO settings (key, value) VALUES ('currency', ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (currency,),
        )


def _format_timestamp(timestamp: float | None) -> str | None:
    if timestamp is None:
        return None
    return datetime.fromtimestamp(timestamp).isoformat(timespec="seconds")


def _get_database_snapshot_path(db_path: str) -> str:
    """Create a temporary consistent SQLite backup for download."""
    tmp = tempfile.NamedTemporaryFile(prefix="financial-tracker-", suffix=".db", delete=False)
    tmp_path = tmp.name
    tmp.close()

    with sqlite3.connect(db_path) as source_conn, sqlite3.connect(tmp_path) as backup_conn:
        source_conn.backup(backup_conn)

    return tmp_path


def _cleanup_snapshot_file(path: str, attempts: int = 5, delay_seconds: float = 0.25) -> None:
    """Remove the temporary snapshot file, tolerating short-lived Windows file locks."""
    for _ in range(max(1, attempts)):
        try:
            if not os.path.exists(path):
                return
            os.remove(path)
            return
        except PermissionError:
            time.sleep(delay_seconds)
        except FileNotFoundError:
            return


def _cleanup_files(*paths: str) -> None:
    for path in paths:
        _cleanup_snapshot_file(path)


def _sanitize_sheet_name(name: str, existing: set[str]) -> str:
    cleaned = ''.join('_' if ch in '\\/*?:[]' else ch for ch in name).strip() or 'Sheet'
    cleaned = cleaned[:31] or 'Sheet'
    candidate = cleaned
    counter = 2
    while candidate in existing:
        suffix = f'_{counter}'
        candidate = f"{cleaned[: max(0, 31 - len(suffix))]}{suffix}" or f"Sheet{suffix}"
        counter += 1
    existing.add(candidate)
    return candidate


def _get_display_column(conn: sqlite3.Connection, table_name: str) -> str | None:
    columns = conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
    for column in columns:
        name = column[1]
        if name.lower() != 'id':
            return name
    return None


def _get_table_names(conn: sqlite3.Connection) -> list[str]:
    rows = conn.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table' AND name NOT LIKE 'sqlite_%'
        ORDER BY name
        """
    ).fetchall()
    return [row[0] for row in rows]


def _group_foreign_keys(conn: sqlite3.Connection, table_name: str) -> list[dict]:
    rows = conn.execute(f'PRAGMA foreign_key_list("{table_name}")').fetchall()
    groups: dict[int, dict] = {}

    for row in rows:
        fk_id = row[0]
        groups.setdefault(
            fk_id,
            {
                "table": row[2],
                "local_columns": [],
                "remote_columns": [],
            },
        )
        groups[fk_id]["local_columns"].append(row[3])
        groups[fk_id]["remote_columns"].append(row[4])

    return list(groups.values())


def _build_foreign_key_maps(conn: sqlite3.Connection, table_name: str) -> dict[str, dict]:
    foreign_keys = _group_foreign_keys(conn, table_name)
    lookup: dict[str, dict] = {}

    for fk in foreign_keys:
        display_column = _get_display_column(conn, fk["table"])
        if not display_column:
            continue

        quoted_remote_columns = ', '.join(f'"{column}"' for column in fk["remote_columns"])
        query = f'SELECT {quoted_remote_columns}, "{display_column}" FROM "{fk["table"]}"'
        rows = conn.execute(query).fetchall()
        label_map = {tuple(row[:-1]): row[-1] for row in rows}

        for local_column in fk["local_columns"]:
            lookup[local_column] = {
                "local_columns": fk["local_columns"],
                "label_column": f"{local_column}_label",
                "label_map": label_map,
            }

    return lookup


def _is_money_column(column_name: str) -> bool:
    normalized = column_name.lower()
    return normalized == "value" or "balance" in normalized or normalized.endswith("_amount")


def _format_export_value(column_name: str, value):
    if value is None:
        return None
    if _is_money_column(column_name) and isinstance(value, int | float):
        return Decimal(value) / Decimal("100")
    return value


def _write_table_sheet(workbook: Workbook, conn: sqlite3.Connection, table_name: str, existing_sheet_names: set[str]) -> None:
    sheet = workbook.create_sheet(title=_sanitize_sheet_name(table_name, existing_sheet_names))
    columns = conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
    column_names = [column[1] for column in columns]
    fk_lookup = _build_foreign_key_maps(conn, table_name)

    headers: list[str] = []
    for column_name in column_names:
        headers.append(column_name)
        if column_name in fk_lookup:
            headers.append(fk_lookup[column_name]["label_column"])

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    rows = conn.execute(f'SELECT * FROM "{table_name}"').fetchall()
    for row in rows:
        row_dict = dict(zip(column_names, row))
        excel_row = []
        for column_name in column_names:
            formatted_value = _format_export_value(column_name, row_dict[column_name])
            excel_row.append(formatted_value)

            fk_info = fk_lookup.get(column_name)
            if fk_info:
                key = tuple(row_dict[local_name] for local_name in fk_info["local_columns"])
                label_value = None if any(part is None for part in key) else fk_info["label_map"].get(key)
                excel_row.append(label_value)

        sheet.append(excel_row)

    for column_index, header in enumerate(headers, start=1):
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in sheet[column_letter]
        )
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)

        if _is_money_column(header):
            for row_index in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                if isinstance(cell.value, Decimal | int | float):
                    cell.number_format = '0.00'


def _export_database_to_excel(snapshot_path: str, output_path: str) -> None:
    with sqlite3.connect(snapshot_path) as conn:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        used_sheet_names: set[str] = set()
        for table_name in _get_table_names(conn):
            _write_table_sheet(workbook, conn, table_name, used_sheet_names)

        workbook.save(output_path)


def _get_excel_export_path(db_path: str) -> tuple[str, str]:
    snapshot_path = _get_database_snapshot_path(db_path)
    filename_base = os.path.splitext(os.path.basename(db_path))[0]
    tmp = tempfile.NamedTemporaryFile(prefix=f"{filename_base}-", suffix=".xlsx", delete=False)
    output_path = tmp.name
    tmp.close()

    _export_database_to_excel(snapshot_path, output_path)
    return snapshot_path, output_path


def _read_database_info() -> dict:
    db_path = get_current_db_path()
    latest_backup_ts = get_latest_backup_time(db_path)
    stat = os.stat(db_path)

    return {
        "path": db_path,
        "filename": os.path.basename(db_path),
        "size_bytes": stat.st_size,
        "last_modified": _format_timestamp(stat.st_mtime),
        "backup_directory": BACKUP_DIR,
        "last_backup": _format_timestamp(latest_backup_ts),
    }


# ── Models ────────────────────────────────────────────────────

class DatabaseInfo(BaseModel):
    path: str
    filename: str
    size_bytes: int
    last_modified: str | None
    backup_directory: str
    last_backup: str | None

class AppConfigResponse(BaseModel):
    currency: str
    database: DatabaseInfo


class AppConfigPatch(BaseModel):
    currency: str


# ── Routes ───────────────────────────────────────────────────

@router.get("", response_model=AppConfigResponse)
def get_app_config():
    """Return the current app configuration (currency)."""
    return {
        "currency": _read_currency(),
        "database": _read_database_info(),
    }


@router.patch("", response_model=AppConfigResponse)
def patch_app_config(body: AppConfigPatch):
    """Update persisted application settings."""
    _write_currency(body.currency.lower())
    return {
        "currency": _read_currency(),
        "database": _read_database_info(),
    }


@router.get("/database/download")
def download_database():
    """Download a consistent snapshot of the active SQLite database."""
    db_path = get_current_db_path()
    snapshot_path = _get_database_snapshot_path(db_path)
    filename = os.path.basename(db_path)

    return FileResponse(
        snapshot_path,
        media_type="application/x-sqlite3",
        filename=filename,
        background=BackgroundTask(_cleanup_snapshot_file, snapshot_path),
    )


@router.get("/database/export-excel")
def export_database_excel():
    """Download all tables from the active SQLite database as an Excel workbook."""
    db_path = get_current_db_path()
    snapshot_path, output_path = _get_excel_export_path(db_path)
    filename = f"{os.path.splitext(os.path.basename(db_path))[0]}.xlsx"

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        background=BackgroundTask(_cleanup_files, snapshot_path, output_path),
    )
